#!/usr/bin/env python3
"""Generate ALAMEIN_CHECKLIST.xlsx — 5 Data Products, 15 tables, ~150 data rows."""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import os

wb = Workbook()

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
PHASE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
PHASE_FONT = Font(name="Calibri", size=12, color="1F4E79", bold=True)
SECTION_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=11, color="375623", bold=True)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SCOPE_FILLS = {1: RED_FILL, 2: YELLOW_FILL, 3: GREEN_FILL}
SEV_FILLS = {"error": RED_FILL, "warn": YELLOW_FILL, "info": GREEN_FILL}

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = THIN_BORDER

def style_row(ws, row, cols, font=None):
    f = font or BODY_FONT
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = f; cell.alignment = WRAP; cell.border = THIN_BORDER

def phase_row(ws, row, text, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = PHASE_FONT; c.fill = PHASE_FILL; c.alignment = WRAP; c.border = THIN_BORDER

def section_row(ws, row, text, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = SECTION_FONT; c.fill = SECTION_FILL; c.alignment = WRAP; c.border = THIN_BORDER

def add_status_dv(ws, col_letter, start_row, end_row):
    dv = DataValidation(type="list", formula1='"☐ Pending,✅ Passed,❌ Failed,⚠️ Skipped"', allow_blank=True)
    dv.error = "Pick from dropdown"; dv.errorTitle = "Invalid"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")

# ═══ TAB 1: MASTER CHECKLIST ═══
ws = wb.active
ws.title = "Checklist"
COLS = 6
for c, w in zip(range(1,7), [5,22,55,35,40,14]):
    ws.column_dimensions[get_column_letter(c)].width = w
for i, h in enumerate(["#","Phase / Section","Step — What to do","URL / Location","Expected Result","Status"], 1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, COLS)
add_status_dv(ws, 'F', 2, 300)

CL = [
    ("PHASE 1 — FOUNDATION (login as ahmed / AdminPa_132)",None,None,None,None,None),
    (None,"1.1 Login","1.1","Open browser, go to login page","http://localhost:5179/login","Login form shown. Sign in as ahmed."),
    (None,"1.1 Login","1.2","Verify landing page","http://localhost:5179/","Platform home with app cards shown."),
    (None,"1.2 Org Units","1.3","Go to Org Units page","http://localhost:5179/admin/org-units","AAST + 5 departments visible."),
    (None,"1.2 Org Units","1.4","VERIFY Alamein Campus exists","http://localhost:5179/admin/org-units","Alamein Campus visible as child of AAST."),
    (None,"1.2 Org Units","1.5","VERIFY 5 departments under Alamein","http://localhost:5179/admin/org-units","Medicine, Financial Affairs, Transportation, Hotels, Hospital all visible."),
    (None,"1.3 Users","1.6","Go to Users page","http://localhost:5179/admin/users","ahmed + 5 alamein.* users."),
    (None,"1.3 Users","1.7","Verify alamein.admin","http://localhost:5179/admin/users","Active."),
    (None,"1.3 Users","1.8","Verify alamein.medical","http://localhost:5179/admin/users","Active."),
    (None,"1.3 Users","1.9","Verify alamein.transport","http://localhost:5179/admin/users","Active."),
    (None,"1.3 Users","1.10","Verify alamein.finance","http://localhost:5179/admin/users","Active."),
    (None,"1.3 Users","1.11","Verify alamein.hotels","http://localhost:5179/admin/users","Active."),
    (None,"1.4 Data Products","1.12","Go to Catalog → Data Products","http://localhost:5179/catalog/products","'New Data Product' visible (admin only)."),
    (None,"1.4 Data Products","1.13","Create DP1: Medicine Carbon (2 tables)","http://localhost:5179/catalog/products","Name=Medicine Carbon, Org Unit=College of Medicine."),
    (None,"1.4 Data Products","1.14","Create DP2: Finance Carbon (3 tables)","http://localhost:5179/catalog/products","Name=Finance Carbon, Org Unit=Financial Affairs."),
    (None,"1.4 Data Products","1.15","Create DP3: Transport Carbon (2 tables)","http://localhost:5179/catalog/products","Name=Transport Carbon, Org Unit=Transportation."),
    (None,"1.4 Data Products","1.16","Create DP4: Hotels Carbon (3 tables)","http://localhost:5179/catalog/products","Name=Hotels Carbon, Org Unit=Student Hotels."),
    (None,"1.4 Data Products","1.17","Create DP5: Hospital Carbon (5 tables)","http://localhost:5179/catalog/products","Name=Hospital Carbon, Org Unit=Educational Hospital."),
    (None,"1.5 Tables + Fields","1.18","Go to Schema Admin → Table Manager","http://localhost:5179/schema-admin/table-manager","Table Manager page loads."),
    (None,"1.5 Tables + Fields","1.19","CREATE ALL 15 TABLES inside their DPs (see Tab: Table List)","http://localhost:5179/schema-admin/table-manager","Each table → Fields → add fields."),
    (None,"1.5 Tables + Fields","1.20","DP1: med_electricity (5 fields)","Table Manager → Fields","period_month, building_id, consumption_kwh, meter_id, cost_egp."),
    (None,"1.5 Tables + Fields","1.21","DP1: med_gen_log (5 fields)","Table Manager → Fields","period_month, generator_id, diesel_liters, runtime_hours, purpose."),
    (None,"1.5 Tables + Fields","1.22","DP2: finance_electricity (5 fields)","Table Manager → Fields","Same as med_electricity."),
    (None,"1.5 Tables + Fields","1.23","DP2: office_supplies (5 fields)","Table Manager → Fields","period_month, paper_reams, paper_type, supplier, cost_egp."),
    (None,"1.5 Tables + Fields","1.24","DP2: med_procurement (5 fields)","Table Manager → Fields","period_month, item_name, category, cost_usd, supplier."),
    (None,"1.5 Tables + Fields","1.25","DP3: fleet_fuel_log (6 fields)","Table Manager → Fields","period_month, vehicle_count, gasoline_liters, diesel_liters, total_cost_egp, supplier."),
    (None,"1.5 Tables + Fields","1.26","DP3: staff_travel (6 fields)","Table Manager → Fields","period_month, staff_name, destination, distance_km, flight_class, cost_egp."),
    (None,"1.5 Tables + Fields","1.27","DP4: hotels_electricity (5 fields)","Table Manager → Fields","Same as med_electricity."),
    (None,"1.5 Tables + Fields","1.28","DP4: hotels_chilled_water (4 fields)","Table Manager → Fields","period_month, meter_id, consumption_tr, building_id."),
    (None,"1.5 Tables + Fields","1.29","DP4: hotels_water (4 fields)","Table Manager → Fields","period_month, building_id, consumption_m3, meter_id."),
    (None,"1.5 Tables + Fields","1.30","DP5: hospital_electricity (5 fields)","Table Manager → Fields","Same as med_electricity."),
    (None,"1.5 Tables + Fields","1.31","DP5: hospital_gen_log (5 fields)","Table Manager → Fields","Same as med_gen_log."),
    (None,"1.5 Tables + Fields","1.32","DP5: medical_gas_log (5 fields)","Table Manager → Fields","period_month, gas_type, quantity_kg, department, purpose."),
    (None,"1.5 Tables + Fields","1.33","DP5: hvac_refrigerant_log (5 fields)","Table Manager → Fields","period_month, unit_id, r410a_kg, service_type, technician."),
    (None,"1.5 Tables + Fields","1.34","DP5: hospital_water (4 fields)","Table Manager → Fields","period_month, building_id, consumption_m3, meter_id."),

    ("PHASE 2 — DATA ENTRY (login as each scoped user)",None,None,None,None,None),
    (None,"2.1 RBAC prep","2.0","Assign ScopedRoles to alamein users","http://localhost:5179/admin/access","Each user has dataowners_group for their org unit. alamein.medical has 2 ScopedRoles (Medicine+Hospital)."),
    (None,"2.1 RBAC test","2.0a","Login alamein.transport → My Data","http://localhost:5179/carbon/my-data","See ONLY DP3 (Transport Carbon)."),
    (None,"2.1 RBAC test","2.0b","Login alamein.hotels → My Data","http://localhost:5179/carbon/my-data","See ONLY DP4 (Hotels Carbon)."),
    (None,"2.1 RBAC test","2.0c","Login alamein.medical → My Data","http://localhost:5179/carbon/my-data","See DP1 + DP5 — 2 Data Products, 7 tables."),
    (None,"2.1 RBAC test","2.0d","Login alamein.finance → My Data","http://localhost:5179/carbon/my-data","See ONLY DP2 (Finance Carbon)."),
    (None,"2.1 RBAC test","2.0e","Login alamein.admin → My Data","http://localhost:5179/carbon/my-data","See ALL 5 Data Products."),
    (None,"2.2 alamein.medical","2.1","Enter DP1/med_electricity data (12 rows)","http://localhost:5179/carbon/my-data","12 monthly rows for MED-101 + MED-102."),
    (None,"2.2 alamein.medical","2.2","Enter DP1/med_gen_log data (5 rows)","http://localhost:5179/carbon/my-data","5 generator fuel logs."),
    (None,"2.2 alamein.medical","2.3","Enter DP5/hospital_electricity data (12 rows)","http://localhost:5179/carbon/my-data","12 rows for HOSP-MAIN + HOSP-WING-B."),
    (None,"2.2 alamein.medical","2.4","Enter DP5/hospital_gen_log data (5 rows)","http://localhost:5179/carbon/my-data","5 generator logs."),
    (None,"2.2 alamein.medical","2.5","Enter DP5/medical_gas_log data (6 rows)","http://localhost:5179/carbon/my-data","6 N₂O entries for Surgery."),
    (None,"2.2 alamein.medical","2.6","Enter DP5/hvac_refrigerant_log data (3 rows)","http://localhost:5179/carbon/my-data","3 R-410A events."),
    (None,"2.2 alamein.medical","2.7","Enter DP5/hospital_water data (6 rows)","http://localhost:5179/carbon/my-data","6 water entries for HOSP-MAIN."),
    (None,"2.3 alamein.transport","2.8","Enter DP3/fleet_fuel_log data (12 rows)","http://localhost:5179/carbon/my-data","12 monthly fleet fuel rows."),
    (None,"2.3 alamein.transport","2.9","Enter DP3/staff_travel data (5 rows)","http://localhost:5179/carbon/my-data","5 staff flights."),
    (None,"2.4 alamein.finance","2.10","Enter DP2/finance_electricity data (6 rows)","http://localhost:5179/carbon/my-data","6 monthly for FIN-TOWER."),
    (None,"2.4 alamein.finance","2.11","Enter DP2/office_supplies data (6 rows)","http://localhost:5179/carbon/my-data","6 monthly paper entries."),
    (None,"2.4 alamein.finance","2.12","Enter DP2/med_procurement data (6 rows)","http://localhost:5179/carbon/my-data","6 procurement entries in USD."),
    (None,"2.5 alamein.hotels","2.13","Enter DP4/hotels_electricity data (12 rows)","http://localhost:5179/carbon/my-data","6 SAKAN-A + 6 SAKAN-B."),
    (None,"2.5 alamein.hotels","2.14","Enter DP4/hotels_chilled_water data (6 rows)","http://localhost:5179/carbon/my-data","6 months for SAKAN-A."),
    (None,"2.5 alamein.hotels","2.15","Enter DP4/hotels_water data (12 rows)","http://localhost:5179/carbon/my-data","6 SAKAN-A + 6 SAKAN-B."),

    ("PHASE 3 — DATA TRUST (DQ + Evidence + Emission Factors)",None,None,None,None,None),
    (None,"3.1 Emission Factors","3.1","Carbon Admin → Emission Factors","http://localhost:5179/carbon/admin/factors","Create 11 factors (see Tab: Emission Factors)."),
    (None,"3.1 Emission Factors","3.2","EGY_GRID_2024","http://localhost:5179/carbon/admin/factors","0.4584 kg CO₂e/kWh, Scope 2."),
    (None,"3.1 Emission Factors","3.3","DIESEL_STATIONARY","http://localhost:5179/carbon/admin/factors","2.68 kg CO₂e/L, Scope 1."),
    (None,"3.1 Emission Factors","3.4","GASOLINE_EG","http://localhost:5179/carbon/admin/factors","2.31 kg CO₂e/L, Scope 1."),
    (None,"3.1 Emission Factors","3.5","CHILLED_WATER_EG","http://localhost:5179/carbon/admin/factors","0.052 kg CO₂e/TR, Scope 2."),
    (None,"3.1 Emission Factors","3.6","WATER_EG","http://localhost:5179/carbon/admin/factors","0.344 kg CO₂e/m³, Scope 3."),
    (None,"3.1 Emission Factors","3.7","PAPER_WASTE_EG","http://localhost:5179/carbon/admin/factors","0.82 kg CO₂e/ream, Scope 3."),
    (None,"3.1 Emission Factors","3.8","FLIGHT_SHORT_EG","http://localhost:5179/carbon/admin/factors","0.15 kg CO₂e/km, Scope 3."),
    (None,"3.1 Emission Factors","3.9","PROCUREMENT_GEN","http://localhost:5179/carbon/admin/factors","0.35 kg CO₂e/USD, Scope 3."),
    (None,"3.1 Emission Factors","3.10","N2O_GWP","http://localhost:5179/carbon/admin/factors","273 kg CO₂e/kg (GWP-100), Scope 1."),
    (None,"3.1 Emission Factors","3.11","R410A_LEAK","http://localhost:5179/carbon/admin/factors","2088 kg CO₂e/kg, Scope 1."),
    (None,"3.2 DQ Rules","3.12","Catalog Studio → DQ Rules","http://localhost:5179/catalog/dq-rules","Create 9 DQ rules (see Tab: DQ Rules)."),
    (None,"3.2 DQ Rules","3.13","DP1/med_electricity: consumption_kwh NOT NULL","http://localhost:5179/catalog/dq-rules","error."),
    (None,"3.2 DQ Rules","3.14","DP1/med_electricity: consumption_kwh range 0-50k","http://localhost:5179/catalog/dq-rules","warn."),
    (None,"3.2 DQ Rules","3.15","DP5/hospital_electricity: consumption_kwh NOT NULL","http://localhost:5179/catalog/dq-rules","error."),
    (None,"3.2 DQ Rules","3.16","DP3/fleet_fuel_log: gasoline_liters NOT NULL","http://localhost:5179/catalog/dq-rules","error."),
    (None,"3.2 DQ Rules","3.17","DP3/fleet_fuel_log: vehicle_count range 1-50","http://localhost:5179/catalog/dq-rules","warn."),
    (None,"3.2 DQ Rules","3.18","DP4/hotels_water: consumption_m3 NOT NULL","http://localhost:5179/catalog/dq-rules","error."),
    (None,"3.2 DQ Rules","3.19","DP4/hotels_water: consumption_m3 range 0-2k","http://localhost:5179/catalog/dq-rules","warn."),
    (None,"3.2 DQ Rules","3.20","DP5/medical_gas_log: quantity_kg NOT NULL","http://localhost:5179/catalog/dq-rules","error."),
    (None,"3.2 DQ Rules","3.21","DP2/office_supplies: paper_reams range 0-500","http://localhost:5179/catalog/dq-rules","info."),
    (None,"3.3 Evidence","3.22","Upload evidence to DP1/med_gen_log row","Row detail → Evidence tab","Evidence appears."),
    (None,"3.3 Evidence","3.23","Upload evidence to DP3/fleet_fuel_log row","Row detail → Evidence tab","Evidence appears."),
    (None,"3.3 Evidence","3.24","Upload evidence to DP5/hospital_electricity row","Row detail → Evidence tab","Evidence appears."),
    (None,"3.3 Evidence","3.25","Upload evidence to DP2/med_procurement row","Row detail → Evidence tab","Evidence appears."),
    (None,"3.4 DQ Execution","3.26","Module workspace → Health tab","Module workspace","DQ score calculated."),
    (None,"3.4 DQ Execution","3.27","Row detail → DQ Metrics tab","Row detail page","Per-row DQ rules shown."),

    ("PHASE 4 — CALCULATIONS",None,None,None,None,None),
    (None,"4.1 Calculation Rules","4.1","Carbon → Calculation Rules","http://localhost:5179/carbon/admin/rules","Create 16 rules (see Tab: Calc Rules)."),
    (None,"4.2 Reporting Period","4.2","Create FY 2024 — Alamein","http://localhost:5179/carbon/reporting/periods","Start=2024-01-01, End=2024-12-31, Type=annual, Status=open."),
    (None,"4.3 Run Calculations","4.3","Carbon → Calculations → Execute","http://localhost:5179/carbon/calculations","All 16 rules execute. No errors."),
    (None,"4.3 Run Calculations","4.4","Verify CO₂e on row detail","Row detail → CO₂e chip","CO₂e value shown."),
    (None,"4.3 Run Calculations","4.5","Verify Lineage tab","Row detail → Lineage tab","Factor → Rule → CO₂e chain."),

    ("PHASE 5 — GOVERNANCE & VERIFICATION",None,None,None,None,None),
    (None,"5.1 Governance","5.1","Catalog → Governance Policies","http://localhost:5179/catalog/policies","Create policy: DP5 (Hospital) lock."),
    (None,"5.2 Verification","5.2","Carbon → Verification","http://localhost:5179/carbon/verification","Pending Review shown."),
    (None,"5.2 Verification","5.3","Approve FY 2024 — Alamein","http://localhost:5179/carbon/verification","Status → Verified."),
    (None,"5.2 Verification","5.4","Try editing verified row","Try editing a verified row","Blocked by governance policy."),

    ("PHASE 6 — UI/UX CROSS-CHECK",None,None,None,None,None),
    (None,"6.1 L1 My Data","6.1","Trust tab shows DQ gauge","http://localhost:5179/carbon/my-data","Gauge renders."),
    (None,"6.1 L1 My Data","6.2","Activity tab filter chips","http://localhost:5179/carbon/my-data","Chips functional."),
    (None,"6.2 L2 Workspace","6.3","Health tab — DQ per table","Module workspace","Table quality scores."),
    (None,"6.2 L2 Workspace","6.4","Governance tab — policies","Module workspace","Active policies for DP."),
    (None,"6.3 L3 Data Entry","6.5","Fields tab","Data Entry page","All fields with type labels."),
    (None,"6.3 L3 Data Entry","6.6","Evidence tab","Data Entry page","Uploaded docs listed."),
    (None,"6.4 L4 Row Detail","6.7","DQ Metrics tab","Row detail","Per-row DQ + rule name."),
    (None,"6.4 L4 Row Detail","6.8","Lineage tab","Row detail","Calculation chain."),
    (None,"6.4 L4 Row Detail","6.9","Related tab","Row detail","FK-linked records."),
    (None,"6.5 Navigation","6.10","Breadcrumbs: L1→L2→L3→L4","All levels","Full trail: Home → DP → Table → Row."),
    (None,"6.5 Navigation","6.11","Browser tab titles","All pages","'{Page} — Carbon Platform'."),
    (None,"6.5 Navigation","6.12","Direct URL navigation","Paste deep link","Page loads correctly."),
    (None,"6.6 Gear icon","6.13","Gear on all 4 levels","Top-right of metrics","Show/hide tabs. Persists."),
    (None,"6.7 Data roundtrip","6.14","Create new row","Add Row","Row saved, appears in grid."),
    (None,"6.7 Data roundtrip","6.15","Edit row","Click row","Changes persist."),
    (None,"6.7 Data roundtrip","6.16","Delete row","Delete button","Row deleted. Count decreases."),

    ("PHASE 7 — BUG BASELINE",None,None,None,None,None),
    (None,"7.1 Title","7.1","L3 browser tab title","Data Entry page","'Table Data — Carbon Platform'."),
    (None,"7.2 Row count","7.2","L2 subtitle row count","Module workspace","Matches visible rows."),
    (None,"7.3 Breadcrumb","7.3","L4 breadcrumb names","Row detail","DP + table names, not IDs."),
    (None,"7.4 History tab","7.4","History entries","Row detail → History","User + action + timestamp."),
    (None,"7.5 Dropdowns","7.5","L1 dropdowns open on click","My Data page","Dropdowns expand."),
]

r = 2
for item in CL:
    phase, section, step_num, text, url, expected = item
    if phase:
        phase_row(ws, r, phase, COLS); style_row(ws, r, COLS, BOLD_FONT)
    elif section and not step_num:
        section_row(ws, r, section, COLS)
    else:
        ws.cell(row=r,column=1,value=step_num); ws.cell(row=r,column=2,value=section)
        ws.cell(row=r,column=3,value=text); ws.cell(row=r,column=4,value=url)
        ws.cell(row=r,column=5,value=expected); ws.cell(row=r,column=6,value="☐ Pending")
        style_row(ws, r, COLS)
    r += 1

ws.conditional_formatting.add(f"F2:F{r}", CellIsRule(operator="equal", formula=['"✅ Passed"'], fill=GREEN_FILL))
ws.conditional_formatting.add(f"F2:F{r}", CellIsRule(operator="equal", formula=['"❌ Failed"'], fill=RED_FILL))
ws.conditional_formatting.add(f"F2:F{r}", CellIsRule(operator="equal", formula=['"⚠️ Skipped"'], fill=YELLOW_FILL))
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:F{r-1}"

# ═══ TAB 2: MODULE LIST (5 Data Products) ═══
ws2 = wb.create_sheet("Module List")
COLS2 = 7
for c, w in zip(range(1,8), [5,30,8,28,14,14,28]):
    ws2.column_dimensions[get_column_letter(c)].width = w
for i, h in enumerate(["#","Data Product Name","Scope","Org Unit","Tables","Status","Notes"], 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, COLS2); add_status_dv(ws2, 'F', 2, 10)

modules = [
    ("DP1","Medicine Carbon",1,"College of Medicine","med_electricity, med_gen_log","2 tables. Scope 1: diesel generators. Scope 2: electricity."),
    ("DP2","Finance Carbon",3,"Financial Affairs","finance_electricity, office_supplies, med_procurement","3 tables. Scope 2: electricity. Scope 3: supplies + procurement."),
    ("DP3","Transport Carbon",1,"Transportation","fleet_fuel_log, staff_travel","2 tables. Scope 1: fleet fuel. Scope 3: staff air travel."),
    ("DP4","Hotels Carbon",2,"Student Hotels","hotels_electricity, hotels_chilled_water, hotels_water","3 tables. Scope 2: electricity + chilled water. Scope 3: water."),
    ("DP5","Hospital Carbon",1,"Educational Hospital","hospital_electricity, hospital_gen_log, medical_gas_log, hvac_refrigerant_log, hospital_water","5 tables. Scope 1: generators+gases+HVAC. Scope 2: electricity. Scope 3: water."),
]
for i, (mid,name,scope,org,tables,notes) in enumerate(modules,2):
    ws2.cell(row=i,column=1,value=mid); ws2.cell(row=i,column=2,value=name)
    ws2.cell(row=i,column=3,value=f"Scope {scope} (dominant)"); ws2.cell(row=i,column=4,value=org)
    ws2.cell(row=i,column=5,value=tables); ws2.cell(row=i,column=6,value="☐ Pending")
    ws2.cell(row=i,column=7,value=notes); style_row(ws2,i,COLS2)
    ws2.cell(row=i,column=3).fill = SCOPE_FILLS[scope]
ws2.freeze_panes = "A2"; ws2.auto_filter.ref = f"A1:G{i}"

# ═══ TAB 3: TABLE LIST + FIELDS ═══
ws3 = wb.create_sheet("Table List + Fields")
COLS3 = 8
for c, w in zip(range(1,9), [5,22,32,16,16,10,8,14]):
    ws3.column_dimensions[get_column_letter(c)].width = w
for i, h in enumerate(["#","Data Product","Table Name","Field Name","Field Label","Type","Req","Status"],1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3,1,COLS3); add_status_dv(ws3,'H',2,100)

tables_fields = [
    ("DP1: Medicine Carbon","med_electricity",[("period_month","Period Month","date","✅"),("building_id","Building ID","string","✅"),("consumption_kwh","Consumption (kWh)","number","✅"),("meter_id","Meter ID","string","❌"),("cost_egp","Cost (EGP)","number","❌")]),
    ("DP1: Medicine Carbon","med_gen_log",[("period_month","Period Month","date","✅"),("generator_id","Generator ID","string","✅"),("diesel_liters","Diesel (L)","number","✅"),("runtime_hours","Runtime Hours","number","✅"),("purpose","Purpose","string","❌")]),
    ("DP2: Finance Carbon","finance_electricity",[("period_month","Period Month","date","✅"),("building_id","Building ID","string","✅"),("consumption_kwh","Consumption (kWh)","number","✅"),("meter_id","Meter ID","string","❌"),("cost_egp","Cost (EGP)","number","❌")]),
    ("DP2: Finance Carbon","office_supplies",[("period_month","Period Month","date","✅"),("paper_reams","Paper (Reams)","number","✅"),("paper_type","Paper Type","string","❌"),("supplier","Supplier","string","❌"),("cost_egp","Cost (EGP)","number","❌")]),
    ("DP2: Finance Carbon","med_procurement",[("period_month","Period Month","date","✅"),("item_name","Item Name","string","✅"),("category","Category","string","❌"),("cost_usd","Cost (USD)","number","✅"),("supplier","Supplier","string","❌")]),
    ("DP3: Transport Carbon","fleet_fuel_log",[("period_month","Period Month","date","✅"),("vehicle_count","Vehicle Count","number","✅"),("gasoline_liters","Gasoline (L)","number","✅"),("diesel_liters","Diesel (L)","number","✅"),("total_cost_egp","Total Cost (EGP)","number","❌"),("supplier","Supplier","string","❌")]),
    ("DP3: Transport Carbon","staff_travel",[("period_month","Period Month","date","✅"),("staff_name","Staff Name","string","✅"),("destination","Destination","string","✅"),("distance_km","Distance (km)","number","✅"),("flight_class","Flight Class","string","❌"),("cost_egp","Cost (EGP)","number","❌")]),
    ("DP4: Hotels Carbon","hotels_electricity",[("period_month","Period Month","date","✅"),("building_id","Building ID","string","✅"),("consumption_kwh","Consumption (kWh)","number","✅"),("meter_id","Meter ID","string","❌"),("cost_egp","Cost (EGP)","number","❌")]),
    ("DP4: Hotels Carbon","hotels_chilled_water",[("period_month","Period Month","date","✅"),("meter_id","Meter ID","string","✅"),("consumption_tr","Consumption (TR)","number","✅"),("building_id","Building ID","string","❌")]),
    ("DP4: Hotels Carbon","hotels_water",[("period_month","Period Month","date","✅"),("building_id","Building ID","string","✅"),("consumption_m3","Consumption (m³)","number","✅"),("meter_id","Meter ID","string","❌")]),
    ("DP5: Hospital Carbon","hospital_electricity",[("period_month","Period Month","date","✅"),("building_id","Building ID","string","✅"),("consumption_kwh","Consumption (kWh)","number","✅"),("meter_id","Meter ID","string","❌"),("cost_egp","Cost (EGP)","number","❌")]),
    ("DP5: Hospital Carbon","hospital_gen_log",[("period_month","Period Month","date","✅"),("generator_id","Generator ID","string","✅"),("diesel_liters","Diesel (L)","number","✅"),("runtime_hours","Runtime Hours","number","✅"),("purpose","Purpose","string","❌")]),
    ("DP5: Hospital Carbon","medical_gas_log",[("period_month","Period Month","date","✅"),("gas_type","Gas Type","string","✅"),("quantity_kg","Quantity (kg)","number","✅"),("department","Department","string","❌"),("purpose","Purpose","string","❌")]),
    ("DP5: Hospital Carbon","hvac_refrigerant_log",[("period_month","Period Month","date","✅"),("unit_id","Unit ID","string","✅"),("r410a_kg","R-410A (kg)","number","✅"),("service_type","Service Type","string","❌"),("technician","Technician","string","❌")]),
    ("DP5: Hospital Carbon","hospital_water",[("period_month","Period Month","date","✅"),("building_id","Building ID","string","✅"),("consumption_m3","Consumption (m³)","number","✅"),("meter_id","Meter ID","string","❌")]),
]

r3 = 2
for module_name, table, fields in tables_fields:
    for fi, (fname, flabel, ftype, freq) in enumerate(fields):
        ws3.cell(row=r3,column=2,value=module_name if fi==0 else "")
        ws3.cell(row=r3,column=3,value=table if fi==0 else "")
        ws3.cell(row=r3,column=4,value=fname); ws3.cell(row=r3,column=5,value=flabel)
        ws3.cell(row=r3,column=6,value=ftype); ws3.cell(row=r3,column=7,value=freq)
        ws3.cell(row=r3,column=8,value="☐ Pending"); style_row(ws3,r3,COLS3); r3 += 1
ws3.freeze_panes = "A2"; ws3.auto_filter.ref = f"A1:H{r3-1}"

# ═══ TAB 4: EMISSION FACTORS ═══
ws4 = wb.create_sheet("Emission Factors")
COLS4 = 8
for c, w in zip(range(1,9), [5,28,20,22,12,8,32,14]):
    ws4.column_dimensions[get_column_letter(c)].width = w
for i, h in enumerate(["#","Name","Code","Value (kg CO₂e/unit)","Unit","Scope","Used By Tables","Status"],1):
    ws4.cell(row=1,column=i,value=h)
style_header(ws4,1,COLS4); add_status_dv(ws4,'H',2,20)

factors = [
    ("Egypt Grid 2024","EGY_GRID_2024","0.4584","kWh",2,"med_electricity, finance_electricity, hotels_electricity, hospital_electricity"),
    ("Diesel — Stationary","DIESEL_STATIONARY","2.68","L",1,"med_gen_log, hospital_gen_log"),
    ("Diesel — Mobile","DIESEL_MOBILE_EG","2.68","L",1,"fleet_fuel_log (diesel — optional)"),
    ("Gasoline — Egypt","GASOLINE_EG","2.31","L",1,"fleet_fuel_log (gasoline)"),
    ("Chilled Water — Egypt","CHILLED_WATER_EG","0.052","TR",2,"hotels_chilled_water"),
    ("Water — Egypt","WATER_EG","0.344","m³",3,"hotels_water, hospital_water"),
    ("Paper Waste — Egypt","PAPER_WASTE_EG","0.82","reams",3,"office_supplies"),
    ("Flight Short Haul — Egypt","FLIGHT_SHORT_EG","0.15","km",3,"staff_travel"),
    ("Procurement — General","PROCUREMENT_GEN","0.35","USD",3,"med_procurement"),
    ("N₂O — GWP 100-year","N2O_GWP","273.0","kg",1,"medical_gas_log"),
    ("R-410A Refrigerant Leak","R410A_LEAK","2088.0","kg",1,"hvac_refrigerant_log"),
]
for i,(name,code,value,unit,scope,tables) in enumerate(factors,2):
    ws4.cell(row=i,column=1,value=i-1); ws4.cell(row=i,column=2,value=name)
    ws4.cell(row=i,column=3,value=code); ws4.cell(row=i,column=4,value=value)
    ws4.cell(row=i,column=5,value=unit); ws4.cell(row=i,column=6,value=f"Scope {scope}")
    ws4.cell(row=i,column=7,value=tables); ws4.cell(row=i,column=8,value="☐ Pending")
    style_row(ws4,i,COLS4); ws4.cell(row=i,column=6).fill = SCOPE_FILLS[scope]
ws4.freeze_panes = "A2"

# ═══ TAB 5: CALC RULES ═══
ws5 = wb.create_sheet("Calc Rules")
COLS5 = 9
for c, w in zip(range(1,10), [5,18,20,22,18,16,8,28,14]):
    ws5.column_dimensions[get_column_letter(c)].width = w
for i, h in enumerate(["#","Data Product","Table","Activity Field","Emission Factor","Date Field","Scope","Expected CO₂e (est.)","Status"],1):
    ws5.cell(row=1,column=i,value=h)
style_header(ws5,1,COLS5); add_status_dv(ws5,'I',2,20)

rules = [
    ("DP1: Medicine","med_electricity","consumption_kwh","EGY_GRID_2024","period_month",2,"~13,780 kg/month avg"),
    ("DP1: Medicine","med_gen_log","diesel_liters","DIESEL_STATIONARY","period_month",1,"~97 kg CO₂e (avg 36 L)"),
    ("DP2: Finance","finance_electricity","consumption_kwh","EGY_GRID_2024","period_month",2,"~20,000 kg/month"),
    ("DP2: Finance","office_supplies","paper_reams","PAPER_WASTE_EG","period_month",3,"~70 kg/month"),
    ("DP2: Finance","med_procurement","cost_usd","PROCUREMENT_GEN","period_month",3,"~4,000 kg/month avg"),
    ("DP3: Transport","fleet_fuel_log","gasoline_liters","GASOLINE_EG","period_month",1,"~4,600 kg/month"),
    ("DP3: Transport","fleet_fuel_log","diesel_liters","DIESEL_MOBILE_EG","period_month",1,"~12,000 kg/month (OPTIONAL)"),
    ("DP3: Transport","staff_travel","distance_km","FLIGHT_SHORT_EG","period_month",3,"~400 kg/trip"),
    ("DP4: Hotels","hotels_electricity","consumption_kwh","EGY_GRID_2024","period_month",2,"~9,100 kg/month"),
    ("DP4: Hotels","hotels_chilled_water","consumption_tr","CHILLED_WATER_EG","period_month",2,"~780 kg/month"),
    ("DP4: Hotels","hotels_water","consumption_m3","WATER_EG","period_month",3,"~215 kg/month avg"),
    ("DP5: Hospital","hospital_electricity","consumption_kwh","EGY_GRID_2024","period_month",2,"~55,000 kg/month"),
    ("DP5: Hospital","hospital_gen_log","diesel_liters","DIESEL_STATIONARY","period_month",1,"~1,870 kg (avg 698 L)"),
    ("DP5: Hospital","medical_gas_log","quantity_kg","N2O_GWP","period_month",1,"~13,500 kg/month (N₂O)"),
    ("DP5: Hospital","hvac_refrigerant_log","r410a_kg","R410A_LEAK","period_month",1,"~6,800 kg CO₂e (avg 3.3 kg)"),
    ("DP5: Hospital","hospital_water","consumption_m3","WATER_EG","period_month",3,"~1,060 kg/month"),
]
for i,(dp,table,field,factor,date_field,scope,est) in enumerate(rules,2):
    ws5.cell(row=i,column=1,value=i-1); ws5.cell(row=i,column=2,value=dp)
    ws5.cell(row=i,column=3,value=table); ws5.cell(row=i,column=4,value=field)
    ws5.cell(row=i,column=5,value=factor); ws5.cell(row=i,column=6,value=date_field)
    ws5.cell(row=i,column=7,value=f"Scope {scope}"); ws5.cell(row=i,column=8,value=est)
    ws5.cell(row=i,column=9,value="☐ Pending"); style_row(ws5,i,COLS5)
    ws5.cell(row=i,column=7).fill = SCOPE_FILLS[scope]
ws5.freeze_panes = "A2"

# ═══ TAB 6: DATA ENTRY ROWS ═══
ws6 = wb.create_sheet("Data Entry Rows")
COLS6 = 10
for c, w in zip(range(1,11), [5,22,18,18,18,18,18,18,16,14]):
    ws6.column_dimensions[get_column_letter(c)].width = w
for i, h in enumerate(["#","DP / Table","period_month","Field 2","Field 3","Field 4","Field 5","Field 6","Who Enters","Status"],1):
    ws6.cell(row=1,column=i,value=h)
style_header(ws6,1,COLS6); add_status_dv(ws6,'J',2,200)

def add_data_rows(ws, start_r, who, rows_data):
    r = start_r
    for module, headers, rows_list in rows_data:
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=COLS6)
        ws.cell(row=r,column=1,value=f"  {module} — Fields: {' → '.join(headers[1:])}")
        ws.cell(row=r,column=1).font = SECTION_FONT; ws.cell(row=r,column=1).fill = SECTION_FILL; ws.cell(row=r,column=1).border = THIN_BORDER
        r += 1
        for row_data in rows_list:
            ws.cell(row=r,column=1,value=""); ws.cell(row=r,column=2,value=module)
            for ci, val in enumerate(row_data,3): ws.cell(row=r,column=ci,value=val)
            ws.cell(row=r,column=9,value=who); ws.cell(row=r,column=10,value="☐ Pending")
            style_row(ws,r,COLS6); r += 1
        r += 1
    return r

r6 = 2
medical_data = [
    ("DP1/med_electricity",["period_month","building_id","consumption_kwh","meter_id","cost_egp"],[
        ("2024-01-01","MED-101",28500,"MTR-MED-101",8150),("2024-02-01","MED-101",26100,"MTR-MED-101",7460),
        ("2024-03-01","MED-101",27300,"MTR-MED-101",7810),("2024-04-01","MED-101",29400,"MTR-MED-101",8400),
        ("2024-05-01","MED-101",31200,"MTR-MED-101",8920),("2024-06-01","MED-101",33500,"MTR-MED-101",9580),
        ("2024-07-01","MED-102",18800,"MTR-MED-102",5370),("2024-08-01","MED-102",17900,"MTR-MED-102",5120),
        ("2024-09-01","MED-102",19200,"MTR-MED-102",5490),("2024-10-01","MED-102",20500,"MTR-MED-102",5860),
        ("2024-11-01","MED-102",19800,"MTR-MED-102",5660),("2024-12-01","MED-102",21300,"MTR-MED-102",6090),
    ]),
    ("DP1/med_gen_log",["period_month","generator_id","diesel_liters","runtime_hours","purpose"],[
        ("2024-03-15","GEN-MED-01",245,14,"Power outage"),("2024-06-20","GEN-MED-01",180,10,"Scheduled test"),
        ("2024-07-10","GEN-MED-01",520,28,"Extended outage"),("2024-09-05","GEN-MED-01",195,11,"Maintenance test"),
        ("2024-11-12","GEN-MED-02",310,18,"Grid failure"),
    ]),
    ("DP5/hospital_electricity",["period_month","building_id","consumption_kwh","meter_id","cost_egp"],[
        ("2024-01-01","HOSP-MAIN",125000,"MTR-HOSP-MAIN",35750),("2024-02-01","HOSP-MAIN",118500,"MTR-HOSP-MAIN",33890),
        ("2024-03-01","HOSP-MAIN",131200,"MTR-HOSP-MAIN",37520),("2024-04-01","HOSP-MAIN",142800,"MTR-HOSP-MAIN",40840),
        ("2024-05-01","HOSP-MAIN",156300,"MTR-HOSP-MAIN",44700),("2024-06-01","HOSP-MAIN",168900,"MTR-HOSP-MAIN",48300),
        ("2024-07-01","HOSP-WING-B",89500,"MTR-HOSP-B",25590),("2024-08-01","HOSP-WING-B",92100,"MTR-HOSP-B",26340),
        ("2024-09-01","HOSP-WING-B",87400,"MTR-HOSP-B",24990),("2024-10-01","HOSP-WING-B",93800,"MTR-HOSP-B",26820),
        ("2024-11-01","HOSP-WING-B",90100,"MTR-HOSP-B",25760),("2024-12-01","HOSP-WING-B",95700,"MTR-HOSP-B",27370),
    ]),
    ("DP5/hospital_gen_log",["period_month","generator_id","diesel_liters","runtime_hours","purpose"],[
        ("2024-02-08","GEN-HOSP-A",890,42,"Power outage — surgery"),("2024-05-15","GEN-HOSP-A",430,22,"Scheduled test"),
        ("2024-08-22","GEN-HOSP-B",1120,55,"Extended blackout"),("2024-10-03","GEN-HOSP-A",380,18,"Grid maintenance"),
        ("2024-12-18","GEN-HOSP-B",670,32,"Storm outage"),
    ]),
    ("DP5/medical_gas_log",["period_month","gas_type","quantity_kg","department","purpose"],[
        ("2024-01-31","N2O",48.5,"Surgery","Anesthesia"),("2024-02-29","N2O",52.1,"Surgery","Anesthesia"),
        ("2024-03-31","N2O",46.8,"Surgery","Anesthesia"),("2024-04-30","N2O",55.3,"Surgery","Anesthesia"),
        ("2024-05-31","N2O",50.7,"Surgery","Anesthesia"),("2024-06-30","N2O",44.2,"Surgery","Anesthesia"),
    ]),
    ("DP5/hvac_refrigerant_log",["period_month","unit_id","r410a_kg","service_type","technician"],[
        ("2024-03-15","AHU-SURG-01",2.8,"Recharge","Eng. Mahmoud"),
        ("2024-07-22","CHILLER-MAIN",5.1,"Leak repair + recharge","Eng. Mahmoud"),
        ("2024-10-10","AHU-ICU-02",1.9,"Recharge","Eng. Samir"),
    ]),
    ("DP5/hospital_water",["period_month","building_id","consumption_m3","meter_id",""],[
        ("2024-01-01","HOSP-MAIN",2850,"WTR-HOSP-MAIN",""),("2024-02-01","HOSP-MAIN",2690,"WTR-HOSP-MAIN",""),
        ("2024-03-01","HOSP-MAIN",2940,"WTR-HOSP-MAIN",""),("2024-04-01","HOSP-MAIN",3120,"WTR-HOSP-MAIN",""),
        ("2024-05-01","HOSP-MAIN",3380,"WTR-HOSP-MAIN",""),("2024-06-01","HOSP-MAIN",3550,"WTR-HOSP-MAIN",""),
    ]),
]
r6 = add_data_rows(ws6, r6, "alamein.medical", medical_data)

transport_data = [
    ("DP3/fleet_fuel_log",["period_month","vehicle_count","gasoline_liters","diesel_liters","total_cost_egp","supplier"],[
        ("2024-01-31",12,1850,4200,78500,"Misr Petroleum"),("2024-02-29",12,1720,3950,73500,"Misr Petroleum"),
        ("2024-03-31",12,1980,4450,83200,"Misr Petroleum"),("2024-04-30",13,2050,4680,87200,"Misr Petroleum"),
        ("2024-05-31",13,1920,4380,81400,"Cooperation"),("2024-06-30",13,1880,4250,79200,"Cooperation"),
        ("2024-07-31",13,2150,4890,91500,"Misr Petroleum"),("2024-08-31",13,2080,4720,88200,"Misr Petroleum"),
        ("2024-09-30",13,1950,4480,83500,"Cooperation"),("2024-10-31",14,2230,5120,95400,"Misr Petroleum"),
        ("2024-11-30",14,2100,4850,90500,"Misr Petroleum"),("2024-12-31",14,1980,4580,85500,"Cooperation"),
    ]),
    ("DP3/staff_travel",["period_month","staff_name","destination","distance_km","flight_class","cost_egp"],[
        ("2024-03-15","Dr. Ahmed Samir","London",3520,"Economy",12450),
        ("2024-05-20","Dr. Layla Hassan","Dubai",2580,"Economy",8950),
        ("2024-06-10","Prof. Khaled Omar","Paris",3210,"Business",28200),
        ("2024-09-05","Dr. Noha Ibrahim","Riyadh",1620,"Economy",6200),
        ("2024-11-18","Dr. Ahmed Samir","Berlin",2950,"Economy",11200),
    ]),
]
r6 = add_data_rows(ws6, r6, "alamein.transport", transport_data)

finance_data = [
    ("DP2/finance_electricity",["period_month","building_id","consumption_kwh","meter_id","cost_egp"],[
        ("2024-01-01","FIN-TOWER",42500,"MTR-FIN-01",12150),("2024-02-01","FIN-TOWER",39800,"MTR-FIN-01",11380),
        ("2024-03-01","FIN-TOWER",41200,"MTR-FIN-01",11780),("2024-04-01","FIN-TOWER",43800,"MTR-FIN-01",12520),
        ("2024-05-01","FIN-TOWER",45600,"MTR-FIN-01",13040),("2024-06-01","FIN-TOWER",48200,"MTR-FIN-01",13780),
    ]),
    ("DP2/office_supplies",["period_month","paper_reams","paper_type","supplier","cost_egp"],[
        ("2024-01-15",85,"A4 80gsm","OfficeMax Egypt",4580),("2024-02-15",72,"A4 80gsm","OfficeMax Egypt",3890),
        ("2024-03-15",95,"A4 80gsm","Office Depot",5120),("2024-04-15",68,"A4 80gsm","OfficeMax Egypt",3670),
        ("2024-05-15",110,"A4 80gsm","Office Depot",5930),("2024-06-15",78,"A4 80gsm","OfficeMax Egypt",4210),
    ]),
    ("DP2/med_procurement",["period_month","item_name","category","cost_usd","supplier"],[
        ("2024-01-20","Surgical gloves (10k)","Consumables",2450,"MedEquip Intl"),
        ("2024-02-20","MRI contrast agent","Imaging",8200,"Siemens Health"),
        ("2024-03-20","ICU ventilators x3","Equipment",48500,"Philips Medical"),
        ("2024-04-20","Surgical sutures","Consumables",1850,"MedEquip Intl"),
        ("2024-05-20","X-ray films","Imaging",3200,"Siemens Health"),
        ("2024-06-20","Blood test reagents","Lab",5600,"Roche Diagnostics"),
    ]),
]
r6 = add_data_rows(ws6, r6, "alamein.finance", finance_data)

hotels_data = [
    ("DP4/hotels_electricity",["period_month","building_id","consumption_kwh","meter_id","cost_egp"],[
        ("2024-01-01","SAKAN-A",18500,"MTR-SAK-A",5290),("2024-02-01","SAKAN-A",17200,"MTR-SAK-A",4920),
        ("2024-03-01","SAKAN-A",19800,"MTR-SAK-A",5660),("2024-04-01","SAKAN-A",21500,"MTR-SAK-A",6150),
        ("2024-05-01","SAKAN-A",23200,"MTR-SAK-A",6630),("2024-06-01","SAKAN-A",24800,"MTR-SAK-A",7090),
        ("2024-01-01","SAKAN-B",15500,"MTR-SAK-B",4430),("2024-02-01","SAKAN-B",14200,"MTR-SAK-B",4060),
        ("2024-03-01","SAKAN-B",16300,"MTR-SAK-B",4660),("2024-04-01","SAKAN-B",17800,"MTR-SAK-B",5090),
        ("2024-05-01","SAKAN-B",19100,"MTR-SAK-B",5460),("2024-06-01","SAKAN-B",20500,"MTR-SAK-B",5860),
    ]),
    ("DP4/hotels_chilled_water",["period_month","meter_id","consumption_tr","building_id",""],[
        ("2024-01-01","CH-SAK-A",12500,"SAKAN-A",""),("2024-02-01","CH-SAK-A",11800,"SAKAN-A",""),
        ("2024-03-01","CH-SAK-A",13200,"SAKAN-A",""),("2024-04-01","CH-SAK-A",14800,"SAKAN-A",""),
        ("2024-05-01","CH-SAK-A",16500,"SAKAN-A",""),("2024-06-01","CH-SAK-A",18200,"SAKAN-A",""),
    ]),
    ("DP4/hotels_water",["period_month","building_id","consumption_m3","meter_id",""],[
        ("2024-01-01","SAKAN-A",620,"WTR-SAK-A",""),("2024-02-01","SAKAN-A",585,"WTR-SAK-A",""),
        ("2024-03-01","SAKAN-A",650,"WTR-SAK-A",""),("2024-04-01","SAKAN-A",710,"WTR-SAK-A",""),
        ("2024-05-01","SAKAN-A",780,"WTR-SAK-A",""),("2024-06-01","SAKAN-A",840,"WTR-SAK-A",""),
        ("2024-01-01","SAKAN-B",510,"WTR-SAK-B",""),("2024-02-01","SAKAN-B",480,"WTR-SAK-B",""),
        ("2024-03-01","SAKAN-B",535,"WTR-SAK-B",""),("2024-04-01","SAKAN-B",590,"WTR-SAK-B",""),
        ("2024-05-01","SAKAN-B",645,"WTR-SAK-B",""),("2024-06-01","SAKAN-B",700,"WTR-SAK-B",""),
    ]),
]
r6 = add_data_rows(ws6, r6, "alamein.hotels", hotels_data)
ws6.freeze_panes = "A2"

# ═══ TAB 7: DQ RULES ═══
ws7 = wb.create_sheet("DQ Rules")
COLS7 = 9
for c, w in zip(range(1,10), [5,18,20,14,25,10,30,10,14]):
    ws7.column_dimensions[get_column_letter(c)].width = w
for i, h in enumerate(["#","Data Product","Table","Field","Rule Type","Severity","Params","Scope","Status"],1):
    ws7.cell(row=1,column=i,value=h)
style_header(ws7,1,COLS7); add_status_dv(ws7,'I',2,15)

dq_rules = [
    ("DP1: Medicine","med_electricity","consumption_kwh","not_null","error","{ }",2),
    ("DP1: Medicine","med_electricity","consumption_kwh","range","warn","{min: 0, max: 50000}",2),
    ("DP5: Hospital","hospital_electricity","consumption_kwh","not_null","error","{ }",2),
    ("DP3: Transport","fleet_fuel_log","gasoline_liters","not_null","error","{ }",1),
    ("DP3: Transport","fleet_fuel_log","vehicle_count","range","warn","{min: 1, max: 50}",1),
    ("DP4: Hotels","hotels_water","consumption_m3","not_null","error","{ }",3),
    ("DP4: Hotels","hotels_water","consumption_m3","range","warn","{min: 0, max: 2000}",3),
    ("DP5: Hospital","medical_gas_log","quantity_kg","not_null","error","{ }",1),
    ("DP2: Finance","office_supplies","paper_reams","range","info","{min: 0, max: 500}",3),
]
for i,(dp,table,field,rtype,sev,params,scope) in enumerate(dq_rules,2):
    ws7.cell(row=i,column=1,value=i-1); ws7.cell(row=i,column=2,value=dp)
    ws7.cell(row=i,column=3,value=table); ws7.cell(row=i,column=4,value=field)
    ws7.cell(row=i,column=5,value=rtype); ws7.cell(row=i,column=6,value=sev)
    ws7.cell(row=i,column=7,value=params); ws7.cell(row=i,column=8,value=f"Scope {scope}")
    ws7.cell(row=i,column=9,value="☐ Pending"); style_row(ws7,i,COLS7)
    ws7.cell(row=i,column=6).fill = SEV_FILLS[sev]; ws7.cell(row=i,column=8).fill = SCOPE_FILLS[scope]
ws7.freeze_panes = "A2"

# ═══ TAB 8: USERS & URLS ═══
ws8 = wb.create_sheet("Users & URLs")
for c, w in zip(range(1,6), [24,18,32,40,14]):
    ws8.column_dimensions[get_column_letter(c)].width = w
section_row(ws8,1,"USER ACCOUNTS",5); style_row(ws8,1,5)
for i,h in enumerate(["Username","Password","Role","Org Unit Scope / Data Products","Status"],1):
    ws8.cell(row=2,column=i,value=h)
style_header(ws8,2,5)

users = [
    ("ahmed","AdminPa_132","Super Admin","All (global) — all 5 DPs"),
    ("alamein.admin","Alamein_2026","Carbon Domain Lead","All 5 DPs (DP1-DP5)"),
    ("alamein.medical","Alamein_2026","Data Owner","DP1: Medicine Carbon + DP5: Hospital Carbon (7 tables)"),
    ("alamein.transport","Alamein_2026","Data Owner","DP3: Transport Carbon only (2 tables)"),
    ("alamein.finance","Alamein_2026","Data Owner","DP2: Finance Carbon only (3 tables)"),
    ("alamein.hotels","Alamein_2026","Data Owner","DP4: Hotels Carbon only (3 tables)"),
]
for i,(uname,pw,role,scope) in enumerate(users,3):
    ws8.cell(row=i,column=1,value=uname); ws8.cell(row=i,column=2,value=pw)
    ws8.cell(row=i,column=3,value=role); ws8.cell(row=i,column=4,value=scope)
    ws8.cell(row=i,column=5,value="☐ Logged in"); style_row(ws8,i,5)

r8 = len(users)+4
section_row(ws8,r8,"KEY URLS",3); style_row(ws8,r8,3); r8+=1
for i,h in enumerate(["Page","URL","Notes"],1): ws8.cell(row=r8,column=i,value=h)
style_header(ws8,r8,3); r8+=1

urls = [
    ("Platform Home","http://localhost:5179/","App cards, nav entry point"),
    ("Login","http://localhost:5179/login","Sign in/out"),
    ("My Data (L1)","http://localhost:5179/carbon/my-data","DPs for scoped users"),
    ("Module Workspace (L2)","http://localhost:5179/carbon/my-data/{moduleId}","Per-DP tables + health"),
    ("Data Entry (L3)","http://localhost:5179/dataschema/entry/{moduleId}/{tableId}","Table data grid"),
    ("Row Detail (L4)","http://localhost:5179/dataschema/row/{tableId}/{rowId}","DQ + lineage"),
    ("Data Products","http://localhost:5179/catalog/products","Create/edit Data Products"),
    ("Table Manager","http://localhost:5179/schema-admin/table-manager","Create tables + fields"),
    ("Org Units","http://localhost:5179/admin/org-units","Create/edit org tree"),
    ("Users","http://localhost:5179/admin/users","Create/edit users"),
    ("Access Control","http://localhost:5179/admin/access","Assign ScopedRoles"),
    ("Emission Factors","http://localhost:5179/carbon/admin/factors","Create/edit factors"),
    ("Calculation Rules","http://localhost:5179/carbon/admin/rules","Bind fields to factors"),
    ("DQ Rules","http://localhost:5179/catalog/dq-rules","Create quality rules"),
    ("Governance Policies","http://localhost:5179/catalog/policies","Delete/edit guards"),
    ("Calculations","http://localhost:5179/carbon/calculations","Run/execute calculations"),
    ("Verification","http://localhost:5179/carbon/verification","Review/approve periods"),
    ("Reporting Periods","http://localhost:5179/carbon/reporting/periods","Create FY periods"),
]
for i,(page,url,notes) in enumerate(urls,r8):
    ws8.cell(row=i,column=1,value=page); ws8.cell(row=i,column=2,value=url)
    ws8.cell(row=i,column=3,value=notes); style_row(ws8,i,3)
ws8.freeze_panes = "A3"

# ═══ SAVE ═══
output_path = os.path.join(os.path.dirname(__file__), "ALAMEIN_CHECKLIST.xlsx")
wb.save(output_path)
print(f"✅ Saved: {output_path}")
print(f"   Tabs: {wb.sheetnames}")
print(f"   8 sheets — 5 Data Products, 15 tables, 16 calc rules, 9 DQ rules, ~150 data rows.")
