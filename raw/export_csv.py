#!/usr/bin/env python3
"""Export AASTMT raw carbon data (Excel + PDF-derived) into clean, seed-ready CSVs.

Sources:
  - raw/Smart_ AASTMT Carbon Emmission_07-07-2025_Magdy (1).xlsx  (Smart Village, FY 2023-24)
  - raw/البصمة الكربونية.pdf  (scanned; Arabic translated manually -> hardcoded below)

Output: raw/csv/*.csv  (idempotent; overwrites)
Usage: .venv/bin/python raw/export_csv.py
"""
from __future__ import annotations
import csv
import os
from datetime import date

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "csv")
XLSX = os.path.join(
    HERE,
    "Smart_ AASTMT Carbon Emmission_07-07-2025_Magdy (1).xlsx",
)

os.makedirs(OUT, exist_ok=True)


def write_csv(name: str, header: list[str], rows: list[list]) -> str:
    path = os.path.join(OUT, name)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    return path


def num(v):
    """Normalize a numeric-ish value to float or ''."""
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return round(float(v), 4)
    s = str(v).strip().replace(",", "")
    if s in ("", "-", "Not Applicable", "Pending ?"):
        return ""
    try:
        return round(float(s), 4)
    except ValueError:
        return s


# --------------------------------------------------------------------------
# 1. Campuses (PDF page 1) — AASTMT's 9 branches
# --------------------------------------------------------------------------
campuses = [
    ("Abu Qir", "abqir", "Alexandria", "Abu Qir Campus", True),          # has electricity data (PDF p2)
    ("Giza Dokki", "dokki", "Giza", "Giza - Dokki Branch", False),
    ("Giza Smart Village", "smart-village", "Giza", "Giza - Smart Village Branch", True),  # main data
    ("Cairo Heliopolis", "heliopolis", "Cairo", "Cairo - Heliopolis Branch", False),
    ("New Alamein", "alamein", "Matrouh", "New Alamein City - Alamein Branch", False),
    ("Alexandria Miami", "miami", "Alexandria", "Alexandria - Miami Branch", False),
    ("Syria Latakia", "latakia", "Syria", "Syria - Latakia Branch", False),
    ("Port Said", "port-said", "Port Said", "Port Said - Port Said Branch", False),
    ("Aswan South Valley", "south-valley", "Aswan", "Aswan - South Valley Branch", True),  # data (PDF p4-6,8)
]
write_csv("campuses.csv",
          ["name", "slug", "city", "full_name", "has_data"],
          [[c[0], c[1], c[2], c[3], "true" if c[4] else "false"] for c in campuses])

# --------------------------------------------------------------------------
# 2. Smart Village monthly meter series (Excel "Detailed data" sheet)
# --------------------------------------------------------------------------
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Detailed data"]

elec, water, chill = [], [], []
for r in range(3, ws.max_row + 1):
    m = ws.cell(row=r, column=1).value
    if not isinstance(m, (date,)):
        continue
    iso = m.strftime("%Y-%m-%d")
    e401, e2401, et = (ws.cell(row=r, column=c).value for c in (2, 3, 4))
    w401, w2401, wt = (ws.cell(row=r, column=c).value for c in (7, 8, 9))
    c1, c2, ct = (ws.cell(row=r, column=c).value for c in (12, 13, 14))
    if e401 is not None or e2401 is not None or et is not None:
        elec.append([iso, num(e401), num(e2401), num(et)])
    if w401 is not None or w2401 is not None or wt is not None:
        water.append([iso, num(w401), num(w2401), num(wt)])
    if c1 is not None or c2 is not None or ct is not None:
        chill.append([iso, num(c1), num(c2), num(ct)])

write_csv("smart_village_monthly_electricity.csv",
          ["month", "building_401_kwh", "building_2401_kwh", "total_kwh"], elec)
write_csv("smart_village_monthly_water.csv",
          ["month", "building_401_m3", "building_2401_m3", "total_m3"], water)
write_csv("smart_village_monthly_chilled_water.csv",
          ["month", "meter_2401_1_tr", "meter_2401_2_tr", "total_tr"], chill)

# --------------------------------------------------------------------------
# 3. Smart Village GHG inventory (Excel "2023-2024" sheet)
# --------------------------------------------------------------------------
ws = wb["2023-2024"]


def existence(quantity, exists):
    """Classify existence: present / not_applicable / pending."""
    if quantity in ("Pending ?", "pending"):
        return "pending"
    if exists == "Not Applicable" or (exists == "" and quantity in ("", "-")):
        return "not_applicable"
    return "present"


inv_rows = []
cur_scope, cur_cat = "", ""
for r in range(14, 69):
    c4 = ws.cell(row=r, column=4).value   # scope
    c5 = ws.cell(row=r, column=5).value   # category
    c6 = ws.cell(row=r, column=6).value   # source
    c7 = ws.cell(row=r, column=7).value   # description (ar)
    c8 = ws.cell(row=r, column=8).value   # activity data
    c9 = ws.cell(row=r, column=9).value   # existence
    c10 = ws.cell(row=r, column=10).value  # unit
    c11 = ws.cell(row=r, column=11).value  # quantity
    if c4:
        cur_scope = str(c4).strip()
    if c5:
        cur_cat = str(c5).strip()
    src = str(c6).strip() if c6 else ""
    act = str(c8).strip() if c8 else ""
    unit = str(c10).strip() if c10 else ""
    qty = str(c11).strip() if c11 else ""
    exists = str(c9).strip() if c9 else ""
    desc = str(c7).strip() if c7 else ""
    if not (src or act):
        continue
    inv_rows.append([
        cur_scope, cur_cat, src, act, desc, existence(qty, exists), unit, num(c11),
    ])

write_csv("smart_village_inventory_fy2324.csv",
          ["scope", "category", "source_of_emission", "activity_data",
           "description_ar", "existence", "unit", "quantity"], inv_rows)

# --------------------------------------------------------------------------
# 4. Abu Qir electricity FY 2025-26 (PDF page 2)
# --------------------------------------------------------------------------
abu = [
    ["2025-07-01", 1027668], ["2025-08-01", 993673], ["2025-09-01", 987273],
    ["2025-10-01", 788411], ["2025-11-01", 757030], ["2025-12-01", 639797],
    ["2026-01-01", 483663], ["2026-02-01", 519901], ["2026-03-01", 483481],
    ["2026-04-01", 506693], ["2026-05-01", 871941], ["2026-06-01", ""],  # June blank in source
]
write_csv("abu_qir_monthly_electricity_fy2526.csv",
          ["month", "total_kwh"], abu)

# --------------------------------------------------------------------------
# 5. Refrigerant cylinders FY 2025-26 (PDF page 3) — main campus + external sites
# --------------------------------------------------------------------------
refrigerants = [
    ["Abu Qir", "R22", 25, "AC units / package units / DX systems"],
    ["Abu Qir", "R134a", 6, "Central AC - simulator bldg + Engineering building G"],
    ["Abu Qir", "R404A", 1, "Refrigeration and freezing rooms"],
    ["Abu Qir", "R410A", 5, "VRV/VRF building systems"],
]
write_csv("refrigerants_fy2526.csv",
          ["campus", "refrigerant", "cylinders_count", "notes"], refrigerants)

# --------------------------------------------------------------------------
# 6. South Valley / Aswan branch, Scope 1 & 2 (PDF page 5)
# --------------------------------------------------------------------------
sv_s12 = [
    ["1", "Stationary Combustion", "On-site generators", "Diesel", "m3", 25],
    ["1", "Mobile Combustion", "Owned/leased vehicles", "Gasoline", "L", 3680],
    ["1", "Mobile Combustion", "Owned/leased vehicles", "Diesel", "L", 125500],
    ["1", "Fugitive Emissions", "Refrigeration/AC systems", "R-404A", "kg", 25],
    ["1", "Fugitive Emissions", "Refrigeration/AC systems", "Other", "kg", 80],
    ["2", "Purchased Electricity", "Purchased electricity", "Electricity", "kWh", 700000],
]
write_csv("south_valley_scope12_fy2526.csv",
          ["scope", "category", "source", "activity_data", "unit", "quantity"], sv_s12)

# --------------------------------------------------------------------------
# 7. South Valley / Aswan branch, Scope 3 (PDF page 6)
# --------------------------------------------------------------------------
sv_s3 = [
    ["Consumable Purchased Goods", "Paper", "Ton", 137.5],
    ["Consumable Purchased Goods", "Envelopes", "Unit", 550],
    ["Consumable Purchased Goods", "Ink / cartridge", "Unit", 3500],
    ["Consumable Purchased Goods", "Toner cartridges", "Unit", 30],
    ["Consumable Purchased Goods", "Toner printer", "Unit", 15],
    ["Hygiene Supplies", "Soap", "Unit", 29000],
    ["Hygiene Supplies", "Tissues", "Unit", 38650],
    ["Capital Goods", "Furniture", "Unit", 70],
    ["Capital Goods", "Facilities", "Unit", 100],
    ["Fertilizers", "Urea / ammonium nitrate / etc.", "Unit", 40],
    ["Fuel & Energy Related", "Electricity T&D losses", "Unit", 200400],
    ["Water Usage / Waste", "Annual water consumption", "m3", 250000],
    ["Upstream Transportation", "Rental vehicles fuel", "Unit", 220000],
    ["Upstream Leased Assets", "Rented office buildings - electricity", "kWh", 960],
    ["Upstream Leased Assets", "Rented office buildings - water", "m3", 200],
]
write_csv("south_valley_scope3_fy2526.csv",
          ["category", "activity_data", "unit", "quantity"], sv_s3)

# --------------------------------------------------------------------------
# 8. Fuel consumption report FY 2025-26 (PDF page 8) — Abu Qir (main campus)
# Financial EGP values are captured in the source but omitted here — not needed
# for GHG activity data (liters only). Attributed to Abu Qir per user (diesel
# total 643,548 L ~5x South Valley's Scope-1 diesel → belongs to main campus).
# --------------------------------------------------------------------------
fuel = [
    ["2025-07-01", 10430.5, 361, 48519],
    ["2025-08-01", 12209.92, 229, 49358],
    ["2025-09-01", 10906, 183, 49306],
    ["2025-10-01", 10678.53, 203, 70607],
    ["2025-11-01", 9558.01, 23, 63643],
    ["2025-12-01", 9921.22, 147, 68972],
    ["2026-01-01", 7963, 305, 48635.69],
    ["2026-02-01", 8592.52, 171, 41962],
    ["2026-03-01", 7649, 298, 52574],
    ["2026-04-01", 7579, 341, 47664],
    ["2026-05-01", 9750, 360, 57110],
    ["2026-06-01", 8953, 315, 45198],
]
write_csv("abu_qir_fuel_fy2526.csv",
          ["month", "gasoline_92_l", "gasoline_95_l", "diesel_l"], fuel)

print("CSVs written to", OUT)
for f in sorted(os.listdir(OUT)):
    print("  -", f)
